from io import BytesIO
from datetime import datetime, timedelta

from flask import (
    Blueprint,
    render_template,
    send_file,
    request,
    session,
    flash
)

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill
)
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import (
    getSampleStyleSheet,
    ParagraphStyle
)
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)

from models import (
    db,
    Habitacion,
    Huesped,
    Reserva,
    Bitacora
)


reportes_bp = Blueprint(
    "reportes",
    __name__
)


def registrar_bitacora(
    accion,
    descripcion
):
    registro = Bitacora(
        accion=accion,
        descripcion=descripcion,
        usuario_id=session.get("usuario_id")
    )

    db.session.add(registro)


def convertir_fecha(valor):
    """
    Convierte una fecha recibida como YYYY-MM-DD
    a un objeto date.
    """

    if not valor:
        return None

    try:
        return datetime.strptime(
            valor,
            "%Y-%m-%d"
        ).date()

    except ValueError:
        return None


def obtener_rango_fechas():
    """
    Obtiene el período solicitado mediante parámetros GET.

    Períodos disponibles:
    - todo
    - hoy
    - semana
    - mes
    - personalizado
    """

    periodo = request.args.get(
        "periodo",
        "todo"
    ).strip().lower()

    fecha_inicio_texto = request.args.get(
        "fecha_inicio",
        ""
    ).strip()

    fecha_fin_texto = request.args.get(
        "fecha_fin",
        ""
    ).strip()

    hoy = datetime.today().date()

    fecha_inicio = None
    fecha_fin = None
    descripcion_periodo = "Todo el historial"

    periodos_validos = {
        "todo",
        "hoy",
        "semana",
        "mes",
        "personalizado"
    }

    if periodo not in periodos_validos:
        periodo = "todo"

    if periodo == "hoy":
        fecha_inicio = hoy
        fecha_fin = hoy
        descripcion_periodo = "Hoy"

    elif periodo == "semana":
        fecha_inicio = hoy - timedelta(
            days=hoy.weekday()
        )

        fecha_fin = fecha_inicio + timedelta(
            days=6
        )

        descripcion_periodo = (
            f"Semana del "
            f"{fecha_inicio.strftime('%d/%m/%Y')} "
            f"al {fecha_fin.strftime('%d/%m/%Y')}"
        )

    elif periodo == "mes":
        fecha_inicio = hoy.replace(
            day=1
        )

        if hoy.month == 12:
            primer_dia_siguiente_mes = hoy.replace(
                year=hoy.year + 1,
                month=1,
                day=1
            )

        else:
            primer_dia_siguiente_mes = hoy.replace(
                month=hoy.month + 1,
                day=1
            )

        fecha_fin = (
            primer_dia_siguiente_mes
            - timedelta(days=1)
        )

        descripcion_periodo = hoy.strftime(
            "%B de %Y"
        )

    elif periodo == "personalizado":
        fecha_inicio = convertir_fecha(
            fecha_inicio_texto
        )

        fecha_fin = convertir_fecha(
            fecha_fin_texto
        )

        if fecha_inicio is None or fecha_fin is None:
            flash(
                "Debes seleccionar una fecha inicial y una fecha final válidas.",
                "warning"
            )

            periodo = "todo"
            fecha_inicio = None
            fecha_fin = None
            descripcion_periodo = "Todo el historial"

        elif fecha_inicio > fecha_fin:
            flash(
                "La fecha inicial no puede ser posterior a la fecha final.",
                "warning"
            )

            periodo = "todo"
            fecha_inicio = None
            fecha_fin = None
            descripcion_periodo = "Todo el historial"

        else:
            descripcion_periodo = (
                f"Del {fecha_inicio.strftime('%d/%m/%Y')} "
                f"al {fecha_fin.strftime('%d/%m/%Y')}"
            )

    return {
        "periodo": periodo,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "fecha_inicio_texto": (
            fecha_inicio.isoformat()
            if fecha_inicio
            else ""
        ),
        "fecha_fin_texto": (
            fecha_fin.isoformat()
            if fecha_fin
            else ""
        ),
        "descripcion_periodo": descripcion_periodo
    }


def obtener_reservas_periodo(
    fecha_inicio=None,
    fecha_fin=None
):
    """
    Devuelve reservas finalizadas, filtradas
    por su fecha de salida.
    """

    consulta = Reserva.query.filter(
        Reserva.estado == "Finalizada",
        Reserva.fecha_salida.isnot(None)
    )

    if fecha_inicio is not None:
        consulta = consulta.filter(
            Reserva.fecha_salida >= fecha_inicio
        )

    if fecha_fin is not None:
        consulta = consulta.filter(
            Reserva.fecha_salida <= fecha_fin
        )

    return consulta.order_by(
        Reserva.fecha_salida.desc(),
        Reserva.id.desc()
    ).all()


def calcular_ingresos(
    reservas
):
    return round(
        sum(
            float(reserva.total_pagado or 0)
            for reserva in reservas
        ),
        2
    )


def obtener_datos_reporte(
    fecha_inicio=None,
    fecha_fin=None
):
    # ==========================================
    # ESTADÍSTICAS OPERATIVAS ACTUALES
    # ==========================================

    total_habitaciones = Habitacion.query.count()

    habitaciones_disponibles = Habitacion.query.filter_by(
        estado="Disponible"
    ).count()

    habitaciones_ocupadas = Habitacion.query.filter_by(
        estado="Ocupada"
    ).count()

    habitaciones_mantenimiento = Habitacion.query.filter_by(
        estado="Mantenimiento"
    ).count()

    total_huespedes = Huesped.query.count()

    reservas_activas = Reserva.query.filter_by(
        estado="Activa"
    ).count()

    if total_habitaciones > 0:
        porcentaje_ocupacion = round(
            habitaciones_ocupadas
            / total_habitaciones
            * 100,
            2
        )

    else:
        porcentaje_ocupacion = 0

    # ==========================================
    # RESERVAS E INGRESOS DEL PERÍODO
    # ==========================================

    reservas_periodo = obtener_reservas_periodo(
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin
    )

    reservas_finalizadas = len(
        reservas_periodo
    )

    ingresos_periodo = calcular_ingresos(
        reservas_periodo
    )

    promedio = 0

    if reservas_finalizadas > 0:
        promedio = round(
            ingresos_periodo
            / reservas_finalizadas,
            2
        )

    # ==========================================
    # INGRESOS DE HOY
    # ==========================================

    hoy = datetime.today().date()

    reservas_hoy = obtener_reservas_periodo(
        fecha_inicio=hoy,
        fecha_fin=hoy
    )

    ingresos_hoy = calcular_ingresos(
        reservas_hoy
    )

    # ==========================================
    # INGRESOS DEL MES ACTUAL
    # ==========================================

    primer_dia_mes = hoy.replace(
        day=1
    )

    if hoy.month == 12:
        primer_dia_siguiente_mes = hoy.replace(
            year=hoy.year + 1,
            month=1,
            day=1
        )

    else:
        primer_dia_siguiente_mes = hoy.replace(
            month=hoy.month + 1,
            day=1
        )

    ultimo_dia_mes = (
        primer_dia_siguiente_mes
        - timedelta(days=1)
    )

    reservas_mes = obtener_reservas_periodo(
        fecha_inicio=primer_dia_mes,
        fecha_fin=ultimo_dia_mes
    )

    ingresos_mes = calcular_ingresos(
        reservas_mes
    )

    # El valor mostrado en "Ingresos totales"
    # corresponde al período seleccionado.
    ingresos_totales = ingresos_periodo

    return {
        "total_habitaciones": total_habitaciones,
        "habitaciones_disponibles": habitaciones_disponibles,
        "habitaciones_ocupadas": habitaciones_ocupadas,
        "habitaciones_mantenimiento": habitaciones_mantenimiento,
        "total_huespedes": total_huespedes,
        "reservas_activas": reservas_activas,
        "reservas_finalizadas": reservas_finalizadas,
        "porcentaje_ocupacion": porcentaje_ocupacion,
        "ingresos_totales": ingresos_totales,
        "ingresos_periodo": ingresos_periodo,
        "ingresos_mes": ingresos_mes,
        "ingresos_hoy": ingresos_hoy,
        "promedio": promedio,
        "reservas": reservas_periodo
    }


@reportes_bp.route("/reportes")
def mostrar_reportes():
    filtro = obtener_rango_fechas()

    datos = obtener_datos_reporte(
        fecha_inicio=filtro["fecha_inicio"],
        fecha_fin=filtro["fecha_fin"]
    )

    return render_template(
        "reportes.html",

        total_habitaciones=datos[
            "total_habitaciones"
        ],

        habitaciones_disponibles=datos[
            "habitaciones_disponibles"
        ],

        habitaciones_ocupadas=datos[
            "habitaciones_ocupadas"
        ],

        habitaciones_mantenimiento=datos[
            "habitaciones_mantenimiento"
        ],

        total_huespedes=datos[
            "total_huespedes"
        ],

        reservas_activas=datos[
            "reservas_activas"
        ],

        reservas_finalizadas=datos[
            "reservas_finalizadas"
        ],

        porcentaje_ocupacion=datos[
            "porcentaje_ocupacion"
        ],

        ingresos_totales=datos[
            "ingresos_totales"
        ],

        ingresos_periodo=datos[
            "ingresos_periodo"
        ],

        ingresos_mes=datos[
            "ingresos_mes"
        ],

        ingresos_hoy=datos[
            "ingresos_hoy"
        ],

        promedio=datos[
            "promedio"
        ],

        ultimas_reservas=datos[
            "reservas"
        ][:10],

        periodo=filtro[
            "periodo"
        ],

        fecha_inicio=filtro[
            "fecha_inicio_texto"
        ],

        fecha_fin=filtro[
            "fecha_fin_texto"
        ],

        descripcion_periodo=filtro[
            "descripcion_periodo"
        ]
    )


@reportes_bp.route("/reportes/excel")
def exportar_excel():
    filtro = obtener_rango_fechas()

    datos = obtener_datos_reporte(
        fecha_inicio=filtro["fecha_inicio"],
        fecha_fin=filtro["fecha_fin"]
    )

    libro = Workbook()

    hoja = libro.active
    hoja.title = "Reporte del Hotel"

    relleno_titulo = PatternFill(
        fill_type="solid",
        fgColor="1D4ED8"
    )

    relleno_encabezado = PatternFill(
        fill_type="solid",
        fgColor="1F2937"
    )

    relleno_financiero = PatternFill(
        fill_type="solid",
        fgColor="DCFCE7"
    )

    fuente_blanca = Font(
        color="FFFFFF",
        bold=True
    )

    hoja.merge_cells(
        "A1:G1"
    )

    celda_titulo = hoja["A1"]

    celda_titulo.value = (
        "HOTEL EL DESCANSO"
    )

    celda_titulo.font = Font(
        color="FFFFFF",
        bold=True,
        size=18
    )

    celda_titulo.fill = relleno_titulo

    celda_titulo.alignment = Alignment(
        horizontal="center"
    )

    hoja.merge_cells(
        "A2:G2"
    )

    hoja["A2"] = (
        "Reporte general del sistema hotelero"
    )

    hoja["A2"].alignment = Alignment(
        horizontal="center"
    )

    hoja.merge_cells(
        "A3:G3"
    )

    hoja["A3"] = (
        f'Período: {filtro["descripcion_periodo"]}'
    )

    hoja["A3"].alignment = Alignment(
        horizontal="center"
    )

    resumen = [
        (
            "Total de habitaciones",
            datos["total_habitaciones"]
        ),
        (
            "Habitaciones disponibles",
            datos["habitaciones_disponibles"]
        ),
        (
            "Habitaciones ocupadas",
            datos["habitaciones_ocupadas"]
        ),
        (
            "Habitaciones en mantenimiento",
            datos["habitaciones_mantenimiento"]
        ),
        (
            "Huéspedes registrados",
            datos["total_huespedes"]
        ),
        (
            "Reservas activas",
            datos["reservas_activas"]
        ),
        (
            "Reservas finalizadas del período",
            datos["reservas_finalizadas"]
        ),
        (
            "Porcentaje de ocupación",
            f'{datos["porcentaje_ocupacion"]}%'
        ),
        (
            "Ingresos del período",
            datos["ingresos_periodo"]
        ),
        (
            "Promedio por reserva",
            datos["promedio"]
        )
    ]

    hoja["A5"] = "Indicador"
    hoja["B5"] = "Valor"

    for celda in hoja[5]:
        celda.fill = relleno_encabezado
        celda.font = fuente_blanca
        celda.alignment = Alignment(
            horizontal="center"
        )

    fila = 6

    for indicador, valor in resumen:
        hoja.cell(
            row=fila,
            column=1,
            value=indicador
        )

        celda_valor = hoja.cell(
            row=fila,
            column=2,
            value=valor
        )

        if indicador in {
            "Ingresos del período",
            "Promedio por reserva"
        }:
            celda_valor.number_format = (
                '"Q"#,##0.00'
            )

            celda_valor.fill = (
                relleno_financiero
            )

        fila += 1

    fila += 2

    encabezados = [
        "ID",
        "Huésped",
        "Habitación",
        "Ingreso",
        "Salida",
        "Total pagado",
        "Estado"
    ]

    for columna, encabezado in enumerate(
        encabezados,
        start=1
    ):
        celda = hoja.cell(
            row=fila,
            column=columna,
            value=encabezado
        )

        celda.fill = relleno_encabezado
        celda.font = fuente_blanca
        celda.alignment = Alignment(
            horizontal="center"
        )

    fila += 1

    for reserva in datos["reservas"]:
        fecha_ingreso = (
            reserva.fecha_ingreso.strftime(
                "%d/%m/%Y"
            )
        )

        fecha_salida = (
            reserva.fecha_salida.strftime(
                "%d/%m/%Y"
            )
            if reserva.fecha_salida
            else "Sin registrar"
        )

        valores = [
            reserva.id,
            reserva.huesped.nombre,
            reserva.habitacion.numero,
            fecha_ingreso,
            fecha_salida,
            float(
                reserva.total_pagado or 0
            ),
            reserva.estado
        ]

        for columna, valor in enumerate(
            valores,
            start=1
        ):
            celda = hoja.cell(
                row=fila,
                column=columna,
                value=valor
            )

            if columna == 6:
                celda.number_format = (
                    '"Q"#,##0.00'
                )

        fila += 1

    anchos = {
        1: 10,
        2: 28,
        3: 15,
        4: 16,
        5: 16,
        6: 18,
        7: 16
    }

    for columna, ancho in anchos.items():
        letra = get_column_letter(
            columna
        )

        hoja.column_dimensions[
            letra
        ].width = ancho

    hoja.freeze_panes = "A6"

    archivo = BytesIO()

    libro.save(
        archivo
    )

    archivo.seek(0)

    registrar_bitacora(
        accion="Exportar Excel",
        descripcion=(
            "Se exportó el reporte en Excel. "
            f'Período: {filtro["descripcion_periodo"]}.'
        )
    )

    db.session.commit()

    fecha_archivo = datetime.now().strftime(
        "%Y%m%d_%H%M"
    )

    return send_file(
        archivo,
        as_attachment=True,
        download_name=(
            f"reporte_hotel_{fecha_archivo}.xlsx"
        ),
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )


@reportes_bp.route("/reportes/pdf")
def exportar_pdf():
    filtro = obtener_rango_fechas()

    datos = obtener_datos_reporte(
        fecha_inicio=filtro["fecha_inicio"],
        fecha_fin=filtro["fecha_fin"]
    )

    archivo = BytesIO()

    documento = SimpleDocTemplate(
        archivo,
        pagesize=landscape(letter),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    estilos = getSampleStyleSheet()

    estilo_titulo = ParagraphStyle(
        "TituloHotel",
        parent=estilos["Title"],
        alignment=TA_CENTER,
        textColor=colors.HexColor(
            "#1D4ED8"
        ),
        fontSize=20,
        spaceAfter=8
    )

    estilo_subtitulo = ParagraphStyle(
        "Subtitulo",
        parent=estilos["Normal"],
        alignment=TA_CENTER,
        textColor=colors.HexColor(
            "#64748B"
        ),
        fontSize=10,
        spaceAfter=8
    )

    contenido = [
        Paragraph(
            "Hotel El Descanso",
            estilo_titulo
        ),
        Paragraph(
            "Reporte general del sistema hotelero",
            estilo_subtitulo
        ),
        Paragraph(
            (
                f'Período: '
                f'{filtro["descripcion_periodo"]}'
            ),
            estilo_subtitulo
        ),
        Spacer(
            1,
            12
        )
    ]

    resumen = [
        [
            "Indicador",
            "Valor"
        ],
        [
            "Total de habitaciones",
            str(
                datos["total_habitaciones"]
            )
        ],
        [
            "Habitaciones disponibles",
            str(
                datos[
                    "habitaciones_disponibles"
                ]
            )
        ],
        [
            "Habitaciones ocupadas",
            str(
                datos[
                    "habitaciones_ocupadas"
                ]
            )
        ],
        [
            "Habitaciones en mantenimiento",
            str(
                datos[
                    "habitaciones_mantenimiento"
                ]
            )
        ],
        [
            "Huéspedes registrados",
            str(
                datos["total_huespedes"]
            )
        ],
        [
            "Reservas activas",
            str(
                datos["reservas_activas"]
            )
        ],
        [
            "Reservas finalizadas del período",
            str(
                datos[
                    "reservas_finalizadas"
                ]
            )
        ],
        [
            "Porcentaje de ocupación",
            f'{datos["porcentaje_ocupacion"]}%'
        ],
        [
            "Ingresos del período",
            f'Q{datos["ingresos_periodo"]:,.2f}'
        ],
        [
            "Promedio por reserva",
            f'Q{datos["promedio"]:,.2f}'
        ]
    ]

    tabla_resumen = Table(
        resumen,
        colWidths=[
            4.5 * inch,
            2.0 * inch
        ]
    )

    tabla_resumen.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor(
                    "#1F2937"
                )
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.HexColor(
                    "#CBD5E1"
                )
            ),
            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.HexColor(
                        "#F8FAFC"
                    )
                ]
            ),
            (
                "ALIGN",
                (1, 1),
                (1, -1),
                "CENTER"
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                7
            ),
            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                7
            )
        ])
    )

    contenido.append(
        tabla_resumen
    )

    contenido.append(
        Spacer(
            1,
            22
        )
    )

    contenido.append(
        Paragraph(
            "Reservas finalizadas del período",
            estilos["Heading2"]
        )
    )

    contenido.append(
        Spacer(
            1,
            8
        )
    )

    historial = [
        [
            "ID",
            "Huésped",
            "Habitación",
            "Ingreso",
            "Salida",
            "Total",
            "Estado"
        ]
    ]

    for reserva in datos["reservas"]:
        historial.append([
            str(
                reserva.id
            ),
            reserva.huesped.nombre,
            str(
                reserva.habitacion.numero
            ),
            reserva.fecha_ingreso.strftime(
                "%d/%m/%Y"
            ),
            (
                reserva.fecha_salida.strftime(
                    "%d/%m/%Y"
                )
                if reserva.fecha_salida
                else "Pendiente"
            ),
            (
                f"Q"
                f"{float(reserva.total_pagado or 0):,.2f}"
            ),
            reserva.estado
        ])

    tabla_historial = Table(
        historial,
        repeatRows=1,
        colWidths=[
            0.5 * inch,
            2.1 * inch,
            1.0 * inch,
            1.0 * inch,
            1.0 * inch,
            1.2 * inch,
            1.0 * inch
        ]
    )

    tabla_historial.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor(
                    "#1F2937"
                )
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),
            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.4,
                colors.HexColor(
                    "#CBD5E1"
                )
            ),
            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.HexColor(
                        "#F8FAFC"
                    )
                ]
            ),
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),
            (
                "ALIGN",
                (0, 0),
                (0, -1),
                "CENTER"
            ),
            (
                "ALIGN",
                (2, 0),
                (-1, -1),
                "CENTER"
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                6
            ),
            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                6
            )
        ])
    )

    contenido.append(
        tabla_historial
    )

    documento.build(
        contenido
    )

    archivo.seek(0)

    registrar_bitacora(
        accion="Exportar PDF",
        descripcion=(
            "Se exportó el reporte en PDF. "
            f'Período: {filtro["descripcion_periodo"]}.'
        )
    )

    db.session.commit()

    fecha_archivo = datetime.now().strftime(
        "%Y%m%d_%H%M"
    )

    return send_file(
        archivo,
        as_attachment=True,
        download_name=(
            f"reporte_hotel_{fecha_archivo}.pdf"
        ),
        mimetype="application/pdf"
    )